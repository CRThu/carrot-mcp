"""Excel tools for carrot-mcp-office using openpyxl."""

from __future__ import annotations

import os
import shutil

import openpyxl
from openpyxl.utils import cell as xl_cell
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference
from openpyxl.styles import Font, Alignment

from carrot_mcp_office._mcp import mcp


def _parse_cell_ref(ref: str) -> tuple[int, int]:
    col_letter, row = xl_cell.coordinate_from_string(ref)
    return row, xl_cell.column_index_from_string(col_letter)


def _parse_range(start: str, end: str) -> tuple[int, int, int, int]:
    min_row, min_col = _parse_cell_ref(start)
    max_row, max_col = _parse_cell_ref(end)
    return min_row, min_col, max_row, max_col


def _backup_file(path: str) -> str | None:
    """Create a .bak copy of the file. Returns backup path or None if no original."""
    if not os.path.exists(path):
        return None
    bak_path = path + ".bak"
    shutil.copy2(path, bak_path)
    return bak_path


def _has_data_in_range(ws, min_row: int, min_col: int, max_row: int, max_col: int) -> bool:
    """Check if any cell in the range has non-None value."""
    for row in ws.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col):
        for cell in row:
            if cell.value is not None:
                return True
    return False


@mcp.tool()
def workbook_metadata(path: str) -> dict:
    """Get workbook metadata (sheet names, properties).

    Args:
        path: Absolute path to the .xlsx file.
    """
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        try:
            sheets = wb.sheetnames
            props = wb.properties
            return {
                "status": "ok",
                "path": path,
                "sheets": sheets,
                "title": props.title,
                "creator": props.creator,
                "created": str(props.created) if props.created else None,
                "modified": str(props.modified) if props.modified else None,
            }
        finally:
            wb.close()
    except Exception as e:
        return {"status": "error", "message": str(e)}


@mcp.tool()
def workbook_search(path: str, sheet: str, query: str) -> dict:
    """Search for values in a sheet.

    Args:
        path: Absolute path to the .xlsx file.
        sheet: Sheet name to search in.
        query: String to search for (case-insensitive substring match).
    """
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        try:
            if sheet not in wb.sheetnames:
                return {"status": "error", "message": f"Sheet '{sheet}' not found"}
            ws = wb[sheet]
            results = []
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and query.lower() in str(cell.value).lower():
                        results.append({"cell": cell.coordinate, "value": cell.value})
            return {"status": "ok", "sheet": sheet, "query": query, "results": results, "count": len(results)}
        finally:
            wb.close()
    except Exception as e:
        return {"status": "error", "message": str(e)}


@mcp.tool()
def create_sheet(path: str, name: str, index: int | None = None, backup: bool = False) -> dict:
    """Create a new sheet in the workbook. Creates the workbook if it doesn't exist.

    Args:
        path: Absolute path to the .xlsx file.
        name: Name for the new sheet.
        index: Position (0-based). None appends at end.
        backup: If True, creates a .bak copy before modifying.
    """
    try:
        if backup:
            _backup_file(path)
        if os.path.exists(path):
            wb = openpyxl.load_workbook(path)
        else:
            wb = openpyxl.Workbook()
        try:
            if name in wb.sheetnames:
                return {"status": "error", "message": f"Sheet '{name}' already exists"}
            ws = wb.create_sheet(title=name, index=index)
            wb.save(path)
            return {"status": "ok", "sheet": name, "index": index}
        finally:
            wb.close()
    except Exception as e:
        return {"status": "error", "message": str(e)}


@mcp.tool()
def rename_sheet(path: str, old_name: str, new_name: str, backup: bool = False) -> dict:
    """Rename a sheet.

    Args:
        path: Absolute path to the .xlsx file.
        old_name: Current sheet name.
        new_name: New sheet name.
        backup: If True, creates a .bak copy before modifying.
    """
    try:
        if backup:
            _backup_file(path)
        wb = openpyxl.load_workbook(path)
        try:
            if old_name not in wb.sheetnames:
                return {"status": "error", "message": f"Sheet '{old_name}' not found"}
            if new_name in wb.sheetnames:
                return {"status": "error", "message": f"Sheet '{new_name}' already exists"}
            ws = wb[old_name]
            ws.title = new_name
            wb.save(path)
            return {"status": "ok", "old_name": old_name, "new_name": new_name}
        finally:
            wb.close()
    except Exception as e:
        return {"status": "error", "message": str(e)}


@mcp.tool()
def delete_sheet(path: str, name: str, backup: bool = False) -> dict:
    """Delete a sheet from the workbook.

    Args:
        path: Absolute path to the .xlsx file.
        name: Sheet name to delete.
        backup: If True, creates a .bak copy before deleting.
    """
    try:
        if backup:
            _backup_file(path)
        wb = openpyxl.load_workbook(path)
        try:
            if name not in wb.sheetnames:
                return {"status": "error", "message": f"Sheet '{name}' not found"}
            if len(wb.sheetnames) <= 1:
                return {"status": "error", "message": "Cannot delete the last sheet"}
            wb.remove(wb[name])
            wb.save(path)
            return {"status": "ok", "sheet": name}
        finally:
            wb.close()
    except Exception as e:
        return {"status": "error", "message": str(e)}


@mcp.tool()
def insert_rows(path: str, sheet: str, start: int, count: int = 1, backup: bool = False) -> dict:
    """Insert rows into a sheet.

    Args:
        path: Absolute path to the .xlsx file.
        sheet: Sheet name.
        start: Row number to start inserting at (1-based).
        count: Number of rows to insert.
        backup: If True, creates a .bak copy before modifying.
    """
    try:
        if backup:
            _backup_file(path)
        wb = openpyxl.load_workbook(path)
        try:
            if sheet not in wb.sheetnames:
                return {"status": "error", "message": f"Sheet '{sheet}' not found"}
            ws = wb[sheet]
            ws.insert_rows(start, count)
            wb.save(path)
            return {"status": "ok", "sheet": sheet, "start": start, "count": count}
        finally:
            wb.close()
    except Exception as e:
        return {"status": "error", "message": str(e)}


@mcp.tool()
def delete_rows(path: str, sheet: str, start: int, count: int = 1, backup: bool = False) -> dict:
    """Delete rows from a sheet.

    Args:
        path: Absolute path to the .xlsx file.
        sheet: Sheet name.
        start: Row number to start deleting at (1-based).
        count: Number of rows to delete.
        backup: If True, creates a .bak copy before deleting.
    """
    try:
        if backup:
            _backup_file(path)
        wb = openpyxl.load_workbook(path)
        try:
            if sheet not in wb.sheetnames:
                return {"status": "error", "message": f"Sheet '{sheet}' not found"}
            ws = wb[sheet]
            ws.delete_rows(start, count)
            wb.save(path)
            return {"status": "ok", "sheet": sheet, "start": start, "count": count}
        finally:
            wb.close()
    except Exception as e:
        return {"status": "error", "message": str(e)}


@mcp.tool()
def insert_columns(path: str, sheet: str, start: int, count: int = 1, backup: bool = False) -> dict:
    """Insert columns into a sheet.

    Args:
        path: Absolute path to the .xlsx file.
        sheet: Sheet name.
        start: Column number to start inserting at (1-based).
        count: Number of columns to insert.
        backup: If True, creates a .bak copy before modifying.
    """
    try:
        if backup:
            _backup_file(path)
        wb = openpyxl.load_workbook(path)
        try:
            if sheet not in wb.sheetnames:
                return {"status": "error", "message": f"Sheet '{sheet}' not found"}
            ws = wb[sheet]
            ws.insert_cols(start, count)
            wb.save(path)
            return {"status": "ok", "sheet": sheet, "start": start, "count": count}
        finally:
            wb.close()
    except Exception as e:
        return {"status": "error", "message": str(e)}


@mcp.tool()
def delete_columns(path: str, sheet: str, start: int, count: int = 1, backup: bool = False) -> dict:
    """Delete columns from a sheet.

    Args:
        path: Absolute path to the .xlsx file.
        sheet: Sheet name.
        start: Column number to start deleting at (1-based).
        count: Number of columns to delete.
        backup: If True, creates a .bak copy before deleting.
    """
    try:
        if backup:
            _backup_file(path)
        wb = openpyxl.load_workbook(path)
        try:
            if sheet not in wb.sheetnames:
                return {"status": "error", "message": f"Sheet '{sheet}' not found"}
            ws = wb[sheet]
            ws.delete_cols(start, count)
            wb.save(path)
            return {"status": "ok", "sheet": sheet, "start": start, "count": count}
        finally:
            wb.close()
    except Exception as e:
        return {"status": "error", "message": str(e)}


@mcp.tool()
def read_range(path: str, sheet: str, start: str, end: str | None = None) -> dict:
    """Read cell values from a range.

    Args:
        path: Absolute path to the .xlsx file.
        sheet: Sheet name.
        start: Start cell reference (e.g. "A1").
        end: End cell reference (e.g. "B2"). None reads single cell.
    """
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        try:
            if sheet not in wb.sheetnames:
                return {"status": "error", "message": f"Sheet '{sheet}' not found"}
            ws = wb[sheet]
            if end is None:
                cell = ws[start]
                return {"status": "ok", "sheet": sheet, "cell": start, "value": cell.value}
            min_row, min_col, max_row, max_col = _parse_range(start, end)
            data = []
            for row in ws.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col):
                data.append([cell.value for cell in row])
            return {"status": "ok", "sheet": sheet, "range": f"{start}:{end}", "data": data}
        finally:
            wb.close()
    except Exception as e:
        return {"status": "error", "message": str(e)}


@mcp.tool()
def write_range(
    path: str,
    sheet: str,
    start: str,
    data: list[list],
    overwrite: bool = False,
    backup: bool = False,
) -> dict:
    """Write a 2D array of values to a range.

    Args:
        path: Absolute path to the .xlsx file.
        sheet: Sheet name.
        start: Start cell reference (e.g. "A1").
        data: 2D array of values to write. Formulas should be strings starting with "=".
        overwrite: If False, returns error when target cells contain data. Set True to allow overwrite.
        backup: If True, creates a .bak copy of the file before writing.
    """
    try:
        wb = openpyxl.load_workbook(path)
        try:
            if sheet not in wb.sheetnames:
                return {"status": "error", "message": f"Sheet '{sheet}' not found"}
            ws = wb[sheet]
            min_row, min_col = _parse_cell_ref(start)
            max_row = min_row + len(data) - 1
            max_col = min_col + (len(data[0]) if data else 1) - 1
            if not overwrite and _has_data_in_range(ws, min_row, min_col, max_row, max_col):
                return {
                    "status": "error",
                    "message": f"Target range starting at {start} contains data. Use overwrite=true to replace.",
                }
            if backup:
                _backup_file(path)
            for r_idx, row in enumerate(data):
                for c_idx, value in enumerate(row):
                    ws.cell(row=min_row + r_idx, column=min_col + c_idx, value=value)
            wb.save(path)
            return {"status": "ok", "sheet": sheet, "start": start, "rows": len(data), "cols": len(data[0]) if data else 0}
        finally:
            wb.close()
    except Exception as e:
        return {"status": "error", "message": str(e)}


@mcp.tool()
def copy_range(
    path: str,
    sheet: str,
    source_start: str,
    source_end: str,
    target_start: str,
    overwrite: bool = False,
    backup: bool = False,
) -> dict:
    """Copy a range of cells to another location.

    Args:
        path: Absolute path to the .xlsx file.
        sheet: Sheet name.
        source_start: Source range start cell (e.g. "A1").
        source_end: Source range end cell (e.g. "B2").
        target_start: Target start cell (e.g. "D1").
        overwrite: If False, returns error when target cells contain data. Set True to allow overwrite.
        backup: If True, creates a .bak copy before modifying.
    """
    try:
        wb = openpyxl.load_workbook(path)
        try:
            if sheet not in wb.sheetnames:
                return {"status": "error", "message": f"Sheet '{sheet}' not found"}
            ws = wb[sheet]
            src_min_row, src_min_col, src_max_row, src_max_col = _parse_range(source_start, source_end)
            tgt_row, tgt_col = _parse_cell_ref(target_start)
            src_rows = src_max_row - src_min_row + 1
            src_cols = src_max_col - src_min_col + 1
            if not overwrite and _has_data_in_range(ws, tgt_row, tgt_col, tgt_row + src_rows - 1, tgt_col + src_cols - 1):
                return {
                    "status": "error",
                    "message": f"Target range starting at {target_start} contains data. Use overwrite=true to replace.",
                }
            if backup:
                _backup_file(path)
            for r in range(src_min_row, src_max_row + 1):
                for c in range(src_min_col, src_max_col + 1):
                    src_cell = ws.cell(row=r, column=c)
                    tgt_cell = ws.cell(row=tgt_row + (r - src_min_row), column=tgt_col + (c - src_min_col))
                    tgt_cell.value = src_cell.value
                    if src_cell.has_style:
                        tgt_cell.font = src_cell.font.copy()
                        tgt_cell.alignment = src_cell.alignment.copy()
                        tgt_cell.border = src_cell.border.copy()
                        tgt_cell.fill = src_cell.fill.copy()
                        tgt_cell.number_format = src_cell.number_format
            wb.save(path)
            return {"status": "ok", "sheet": sheet, "source": f"{source_start}:{source_end}", "target": target_start}
        finally:
            wb.close()
    except Exception as e:
        return {"status": "error", "message": str(e)}


@mcp.tool()
def delete_range(path: str, sheet: str, start: str, end: str, backup: bool = False) -> dict:
    """Clear cell contents in a range (keeps formatting).

    Args:
        path: Absolute path to the .xlsx file.
        sheet: Sheet name.
        start: Start cell reference (e.g. "A1").
        end: End cell reference (e.g. "B2").
        backup: If True, creates a .bak copy before clearing.
    """
    try:
        if backup:
            _backup_file(path)
        wb = openpyxl.load_workbook(path)
        try:
            if sheet not in wb.sheetnames:
                return {"status": "error", "message": f"Sheet '{sheet}' not found"}
            ws = wb[sheet]
            min_row, min_col, max_row, max_col = _parse_range(start, end)
            for row in ws.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col):
                for cell in row:
                    cell.value = None
            wb.save(path)
            return {"status": "ok", "sheet": sheet, "range": f"{start}:{end}"}
        finally:
            wb.close()
    except Exception as e:
        return {"status": "error", "message": str(e)}


@mcp.tool()
def read_chart(path: str, sheet: str) -> dict:
    """Read chart information from a sheet.

    Args:
        path: Absolute path to the .xlsx file.
        sheet: Sheet name.
    """
    try:
        wb = openpyxl.load_workbook(path)
        try:
            if sheet not in wb.sheetnames:
                return {"status": "error", "message": f"Sheet '{sheet}' not found"}
            ws = wb[sheet]
            charts = []
            for i, chart in enumerate(ws._charts):
                charts.append({
                    "index": i,
                    "type": type(chart).__name__,
                    "title": chart.title,
                })
            return {"status": "ok", "sheet": sheet, "charts": charts, "count": len(charts)}
        finally:
            wb.close()
    except Exception as e:
        return {"status": "error", "message": str(e)}


@mcp.tool()
def write_chart(
    path: str,
    sheet: str,
    chart_type: str,
    data_range: str,
    target_cell: str,
    title: str = "",
    backup: bool = False,
) -> dict:
    """Create a chart in the sheet.

    Args:
        path: Absolute path to the .xlsx file.
        sheet: Sheet name.
        chart_type: One of "bar", "line", "pie", "scatter".
        data_range: Data range reference (e.g. "A1:B5").
        target_cell: Cell where chart will be placed (e.g. "D1").
        title: Chart title.
        backup: If True, creates a .bak copy before modifying.
    """
    try:
        if backup:
            _backup_file(path)
        wb = openpyxl.load_workbook(path)
        try:
            if sheet not in wb.sheetnames:
                return {"status": "error", "message": f"Sheet '{sheet}' not found"}
            ws = wb[sheet]
            chart_map = {
                "bar": BarChart,
                "line": LineChart,
                "pie": PieChart,
                "scatter": ScatterChart,
            }
            if chart_type not in chart_map:
                return {"status": "error", "message": f"Unknown chart type '{chart_type}'. Use: bar, line, pie, scatter"}
            chart = chart_map[chart_type]()
            chart.title = title
            min_row, min_col, max_row, max_col = _parse_range(data_range.split(":")[0], data_range.split(":")[1])
            data_ref = Reference(ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)
            chart.add_data(data_ref, titles_from_data=True)
            tgt_row, tgt_col = _parse_cell_ref(target_cell)
            ws.add_chart(chart, f"{xl_cell.get_column_letter(tgt_col)}{tgt_row}")
            wb.save(path)
            return {"status": "ok", "sheet": sheet, "chart_type": chart_type, "target": target_cell}
        finally:
            wb.close()
    except Exception as e:
        return {"status": "error", "message": str(e)}


@mcp.tool()
def format_range(
    path: str,
    sheet: str,
    start: str,
    end: str,
    font_name: str | None = None,
    font_size: int | None = None,
    font_color: str | None = None,
    bold: bool | None = None,
    italic: bool | None = None,
    underline: bool | None = None,
    alignment: str | None = None,
    number_format: str | None = None,
    merge: bool = False,
    unmerge: bool = False,
    backup: bool = False,
) -> dict:
    """Format a range of cells.

    Args:
        path: Absolute path to the .xlsx file.
        sheet: Sheet name.
        start: Start cell reference (e.g. "A1").
        end: End cell reference (e.g. "B2").
        font_name: Font family name.
        font_size: Font size in points.
        font_color: Font color as hex string (e.g. "FF0000").
        bold: Bold text.
        italic: Italic text.
        underline: Underline text.
        alignment: Text alignment (left, center, right, justify).
        number_format: Number format string (e.g. "0.00", "yyyy-mm-dd").
        merge: Merge cells in range.
        unmerge: Unmerge cells in range.
        backup: If True, creates a .bak copy before formatting.
    """
    try:
        if backup:
            _backup_file(path)
        wb = openpyxl.load_workbook(path)
        try:
            if sheet not in wb.sheetnames:
                return {"status": "error", "message": f"Sheet '{sheet}' not found"}
            ws = wb[sheet]
            min_row, min_col, max_row, max_col = _parse_range(start, end)
            for row in ws.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col):
                for cell in row:
                    if font_name or font_size or font_color or bold is not None or italic is not None or underline is not None:
                        old = cell.font
                        cell.font = Font(
                            name=font_name or old.name,
                            size=font_size or old.size,
                            color=font_color or old.color,
                            bold=bold if bold is not None else old.bold,
                            italic=italic if italic is not None else old.italic,
                            underline=underline if underline is not None else old.underline,
                        )
                    if alignment:
                        align_map = {"left": "left", "center": "center", "right": "right", "justify": "justify"}
                        if alignment not in align_map:
                            return {"status": "error", "message": f"Invalid alignment '{alignment}'. Use: left, center, right, justify"}
                        cell.alignment = Alignment(horizontal=alignment)
                    if number_format:
                        cell.number_format = number_format
            if merge:
                ws.merge_cells(f"{start}:{end}")
            elif unmerge:
                ws.unmerge_cells(f"{start}:{end}")
            wb.save(path)
            return {"status": "ok", "sheet": sheet, "range": f"{start}:{end}"}
        finally:
            wb.close()
    except Exception as e:
        return {"status": "error", "message": str(e)}
