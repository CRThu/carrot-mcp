"""Tests for Excel tools."""

import os
import shutil
import tempfile

from carrot_mcp_office.excel import (
    workbook_metadata,
    workbook_grep,
    create_sheet,
    rename_sheet,
    delete_sheet,
    insert_rows,
    delete_rows,
    insert_columns,
    delete_columns,
    read_range,
    write_range,
    copy_range,
    delete_range,
    read_chart,
    write_chart,
    format_range,
)


def _cleanup(original_path):
    from carrot_mcp_office.backup import _mirror_path
    mirror = _mirror_path(original_path)
    if mirror.parent.exists():
        shutil.rmtree(mirror.parent, ignore_errors=True)
    d = os.path.dirname(original_path)
    if os.path.exists(d):
        shutil.rmtree(d, ignore_errors=True)


def _xlsx():
    d = tempfile.mkdtemp(prefix="test_office_")
    return os.path.join(d, "test.xlsx")


def test_create_sheet_new_file():
    path = _xlsx()
    try:
        result = create_sheet(path, "Sheet1")
        assert result["status"] == "ok"
        assert result["sheet"] == "Sheet1"
        assert result["version"] >= 1
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_create_sheet_existing_file():
    path = _xlsx()
    try:
        create_sheet(path, "Sheet1")
        result = create_sheet(path, "Sheet2")
        assert result["status"] == "ok"
        meta = workbook_metadata(path)
        assert "Sheet1" in meta["sheets"]
        assert "Sheet2" in meta["sheets"]
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_create_sheet_duplicate_name():
    path = _xlsx()
    try:
        create_sheet(path, "Sheet1")
        result = create_sheet(path, "Sheet1")
        assert result["status"] == "error"
        assert "already exists" in result["message"]
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_workbook_metadata():
    path = _xlsx()
    try:
        create_sheet(path, "Data")
        result = workbook_metadata(path)
        assert result["status"] == "ok"
        assert "Data" in result["sheets"]
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_workbook_metadata_nonexistent():
    result = workbook_metadata("nonexistent.xlsx")
    assert result["status"] == "error"


def test_rename_sheet():
    path = _xlsx()
    try:
        create_sheet(path, "OldName")
        result = rename_sheet(path, "OldName", "NewName")
        assert result["status"] == "ok"
        assert result["version"] >= 1
        meta = workbook_metadata(path)
        assert "NewName" in meta["sheets"]
        assert "OldName" not in meta["sheets"]
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_rename_sheet_not_found():
    path = _xlsx()
    try:
        create_sheet(path, "Sheet1")
        result = rename_sheet(path, "Nonexistent", "NewName")
        assert result["status"] == "error"
        assert "not found" in result["message"]
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_delete_sheet():
    path = _xlsx()
    try:
        create_sheet(path, "Sheet1")
        create_sheet(path, "Sheet2")
        result = delete_sheet(path, "Sheet1")
        assert result["status"] == "ok"
        assert result["version"] >= 1
        meta = workbook_metadata(path)
        assert "Sheet1" not in meta["sheets"]
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_delete_last_sheet():
    path = _xlsx()
    try:
        create_sheet(path, "Sheet1")
        delete_sheet(path, "Sheet")
        result = delete_sheet(path, "Sheet1")
        assert result["status"] == "error"
        assert "last sheet" in result["message"]
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_write_read_range():
    path = _xlsx()
    try:
        create_sheet(path, "Data")
        write_range(path, "Data", "A1", [["Name", "Score"], ["Alice", 95], ["Bob", 87]])
        result = read_range(path, "Data", "A1", "B3")
        assert result["status"] == "ok"
        assert result["data"] == [["Name", "Score"], ["Alice", 95], ["Bob", 87]]
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_write_range_overwrite_protection():
    path = _xlsx()
    try:
        create_sheet(path, "Data")
        write_range(path, "Data", "A1", [["Original"]])
        result = write_range(path, "Data", "A1", [["Overwritten"]])
        assert result["status"] == "error"
        assert "overwrite" in result["message"].lower()
        val = read_range(path, "Data", "A1")
        assert val["value"] == "Original"
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_write_range_overwrite_allowed():
    path = _xlsx()
    try:
        create_sheet(path, "Data")
        write_range(path, "Data", "A1", [["Original"]])
        result = write_range(path, "Data", "A1", [["Overwritten"]], overwrite=True)
        assert result["status"] == "ok"
        assert result["version"] >= 1
        val = read_range(path, "Data", "A1")
        assert val["value"] == "Overwritten"
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_copy_range_overwrite_protection():
    path = _xlsx()
    try:
        create_sheet(path, "Data")
        write_range(path, "Data", "A1", [["Source"]])
        write_range(path, "Data", "C1", [["Target"]])
        result = copy_range(path, "Data", "A1", "A1", "C1")
        assert result["status"] == "error"
        assert "overwrite" in result["message"].lower()
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_copy_range_overwrite_allowed():
    path = _xlsx()
    try:
        create_sheet(path, "Data")
        write_range(path, "Data", "A1", [["Source"]])
        write_range(path, "Data", "C1", [["Target"]])
        result = copy_range(path, "Data", "A1", "A1", "C1", overwrite=True)
        assert result["status"] == "ok"
        assert result["version"] >= 1
        val = read_range(path, "Data", "C1")
        assert val["value"] == "Source"
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_delete_rows():
    path = _xlsx()
    try:
        create_sheet(path, "Data")
        write_range(path, "Data", "A1", [["A"], ["B"], ["C"]])
        delete_rows(path, "Data", 2, 1)
        result = read_range(path, "Data", "A1", "A2")
        assert result["status"] == "ok"
        assert result["data"] == [["A"], ["C"]]
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_read_single_cell():
    path = _xlsx()
    try:
        create_sheet(path, "Data")
        write_range(path, "Data", "A1", [["Hello"]])
        result = read_range(path, "Data", "A1")
        assert result["status"] == "ok"
        assert result["value"] == "Hello"
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_read_range_nonexistent_sheet():
    path = _xlsx()
    try:
        create_sheet(path, "Data")
        result = read_range(path, "Nonexistent", "A1")
        assert result["status"] == "error"
        assert "not found" in result["message"]
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_insert_delete_rows():
    path = _xlsx()
    try:
        create_sheet(path, "Data")
        write_range(path, "Data", "A1", [["A"], ["B"], ["C"]])
        insert_rows(path, "Data", 2, 1)
        result = read_range(path, "Data", "A1", "A4")
        assert result["status"] == "ok"
        assert result["data"] == [["A"], [None], ["B"], ["C"]]
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_insert_delete_columns():
    path = _xlsx()
    try:
        create_sheet(path, "Data")
        write_range(path, "Data", "A1", [["A", "B"]])
        insert_columns(path, "Data", 2, 1)
        result = read_range(path, "Data", "A1", "C1")
        assert result["status"] == "ok"
        assert result["data"] == [["A", None, "B"]]
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_copy_range():
    path = _xlsx()
    try:
        create_sheet(path, "Data")
        write_range(path, "Data", "A1", [["X", "Y"]])
        copy_range(path, "Data", "A1", "B1", "D1")
        result = read_range(path, "Data", "D1", "E1")
        assert result["status"] == "ok"
        assert result["data"] == [["X", "Y"]]
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_delete_range():
    path = _xlsx()
    try:
        create_sheet(path, "Data")
        write_range(path, "Data", "A1", [["X", "Y"]])
        delete_range(path, "Data", "A1", "B1")
        result = read_range(path, "Data", "A1", "B1")
        assert result["status"] == "ok"
        assert result["data"] == [[None, None]]
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_format_range():
    path = _xlsx()
    try:
        create_sheet(path, "Data")
        write_range(path, "Data", "A1", [["Test"]])
        result = format_range(path, "Data", "A1", "A1", bold=True, font_size=14)
        assert result["status"] == "ok"
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_format_range_invalid_alignment():
    path = _xlsx()
    try:
        create_sheet(path, "Data")
        result = format_range(path, "Data", "A1", "A1", alignment="invalid")
        assert result["status"] == "error"
        assert "Invalid alignment" in result["message"]
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_format_range_merge():
    path = _xlsx()
    try:
        create_sheet(path, "Data")
        result = format_range(path, "Data", "A1", "B2", merge=True)
        assert result["status"] == "ok"
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_workbook_grep():
    path = _xlsx()
    try:
        create_sheet(path, "Data")
        write_range(path, "Data", "A1", [["Apple"], ["Banana"], ["Cherry"]])
        result = workbook_grep(path, "Data", "ana")
        assert result["status"] == "ok"
        assert result["count"] == 1
        assert result["results"][0]["cell"] == "A2"
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_workbook_grep_no_results():
    path = _xlsx()
    try:
        create_sheet(path, "Data")
        write_range(path, "Data", "A1", [["Apple"]])
        result = workbook_grep(path, "Data", "xyz")
        assert result["status"] == "ok"
        assert result["count"] == 0
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_workbook_grep_regex():
    path = _xlsx()
    try:
        create_sheet(path, "Data")
        write_range(path, "Data", "A1", [["abc 123"], ["xyz 456"], ["abc 789"]])
        result = workbook_grep(path, "Data", r"abc \d+", regex=True)
        assert result["status"] == "ok"
        assert result["count"] == 2
        assert result["regex"] is True
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_workbook_grep_regex_invalid():
    path = _xlsx()
    try:
        create_sheet(path, "Data")
        write_range(path, "Data", "A1", [["test"]])
        result = workbook_grep(path, "Data", r"[invalid", regex=True)
        assert result["status"] == "error"
        assert "Invalid regex" in result["message"]
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_write_chart():
    path = _xlsx()
    try:
        create_sheet(path, "Data")
        write_range(path, "Data", "A1", [["Month", "Sales"], ["Jan", 100], ["Feb", 200]])
        result = write_chart(path, "Data", "bar", "A1:B3", "D1", title="Sales")
        assert result["status"] == "ok"
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_write_chart_invalid_type():
    path = _xlsx()
    try:
        create_sheet(path, "Data")
        result = write_chart(path, "Data", "invalid", "A1:B3", "D1")
        assert result["status"] == "error"
        assert "Unknown chart type" in result["message"]
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_read_chart():
    path = _xlsx()
    try:
        create_sheet(path, "Data")
        write_range(path, "Data", "A1", [["Month", "Sales"], ["Jan", 100]])
        write_chart(path, "Data", "bar", "A1:B2", "D1", title="Sales")
        result = read_chart(path, "Data")
        assert result["status"] == "ok"
        assert result["count"] == 1
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_format_range_unmerge():
    path = _xlsx()
    try:
        create_sheet(path, "Data")
        format_range(path, "Data", "A1", "B2", merge=True)
        import openpyxl
        wb = openpyxl.load_workbook(path)
        assert len(list(wb["Data"].merged_cells.ranges)) == 1
        wb.close()
        format_range(path, "Data", "A1", "B2", unmerge=True)
        wb = openpyxl.load_workbook(path)
        assert len(list(wb["Data"].merged_cells.ranges)) == 0
        wb.close()
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_format_range_number_format():
    path = _xlsx()
    try:
        create_sheet(path, "Data")
        write_range(path, "Data", "A1", [[1234.56]])
        result = format_range(path, "Data", "A1", "A1", number_format="#,##0.00")
        assert result["status"] == "ok"
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_format_range_font_color():
    path = _xlsx()
    try:
        create_sheet(path, "Data")
        write_range(path, "Data", "A1", [["Red"]])
        result = format_range(path, "Data", "A1", "A1", font_color="FF0000", bold=True)
        assert result["status"] == "ok"
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_write_range_formula():
    path = _xlsx()
    try:
        create_sheet(path, "Data")
        write_range(path, "Data", "A1", [[10, 20, "=A1+B1"]])
        result = read_range(path, "Data", "C1")
        assert result["status"] == "ok"
        assert result["value"] == "=A1+B1"
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_delete_sheet_not_found():
    path = _xlsx()
    try:
        create_sheet(path, "Sheet1")
        result = delete_sheet(path, "Nonexistent")
        assert result["status"] == "error"
        assert "not found" in result["message"]
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_insert_rows_not_found():
    path = _xlsx()
    try:
        create_sheet(path, "Sheet1")
        result = insert_rows(path, "Nonexistent", 1)
        assert result["status"] == "error"
        assert "not found" in result["message"]
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_delete_rows_not_found():
    path = _xlsx()
    try:
        create_sheet(path, "Sheet1")
        result = delete_rows(path, "Nonexistent", 1)
        assert result["status"] == "error"
        assert "not found" in result["message"]
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_insert_columns_not_found():
    path = _xlsx()
    try:
        create_sheet(path, "Sheet1")
        result = insert_columns(path, "Nonexistent", 1)
        assert result["status"] == "error"
        assert "not found" in result["message"]
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)


def test_delete_columns_not_found():
    path = _xlsx()
    try:
        create_sheet(path, "Sheet1")
        result = delete_columns(path, "Nonexistent", 1)
        assert result["status"] == "error"
        assert "not found" in result["message"]
    finally:
        _cleanup(path)
        if os.path.exists(path):
            os.unlink(path)
